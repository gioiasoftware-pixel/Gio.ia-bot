"""
Processore CSV/Excel per inventario
Estrae dati vini da file CSV e Excel
"""
import io
import logging
from typing import List, Dict, Any
from openpyxl import load_workbook
import csv

logger = logging.getLogger(__name__)

class CSVProcessor:
    """Processore CSV/Excel per inventario"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    async def process_file(self, file_content: bytes, file_name: str, file_type: str) -> List[Dict[str, Any]]:
        """
        Elabora un file CSV o Excel
        
        Args:
            file_content: Contenuto binario del file
            file_name: Nome del file
            file_type: Tipo file ('csv' o 'excel')
            
        Returns:
            Lista di vini estratti
        """
        try:
            self.logger.info(f"Elaborazione file {file_type}: {file_name}")
            
            # Leggi il file in base al tipo
            if file_type == 'csv':
                wines = self._process_csv(file_content)
            elif file_type == 'excel':
                wines = self._process_excel(file_content)
            else:
                raise ValueError(f"Tipo file non supportato: {file_type}")
            
            self.logger.info(f"File letto: {len(wines)} vini estratti")
            
            self.logger.info(f"Vini estratti: {len(wines)}")
            return wines
            
        except Exception as e:
            self.logger.error(f"Errore elaborazione file: {e}")
            return []
    
    def _process_csv(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Processa file CSV"""
        wines = []
        csv_reader = csv.DictReader(io.StringIO(file_content.decode('utf-8')))
        
        for row in csv_reader:
            wine_data = self._extract_wine_from_row(row)
            if wine_data:
                wines.append(wine_data)
        
        return wines
    
    def _process_excel(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Processa file Excel"""
        wines = []
        workbook = load_workbook(io.BytesIO(file_content))
        sheet = workbook.active
        
        # Leggi header
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Leggi dati
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            wine_data = self._extract_wine_from_row(row_dict)
            if wine_data:
                wines.append(wine_data)
        
        return wines
    
    def _extract_wine_from_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """
        Estrae dati vino da una riga
        
        Args:
            row: Riga del file
            
        Returns:
            Dati vino o None
        """
        try:
            # Estrai nome
            name = self._get_column_value(row, 'name')
            if not name or str(name).strip() == '':
                return None
            
            # Estrai produttore
            producer = self._get_column_value(row, 'producer')
            if not producer:
                producer = 'Sconosciuto'
            
            # Estrai quantità
            quantity = self._get_column_value(row, 'quantity')
            if not quantity:
                quantity = 1
            else:
                try:
                    quantity = int(float(quantity))
                except (ValueError, TypeError):
                    quantity = 1
            
            # Estrai altri dati
            vintage = self._get_column_value(row, 'vintage')
            wine_type = self._get_column_value(row, 'wine_type')
            region = self._get_column_value(row, 'region')
            price = self._get_column_value(row, 'price')
            min_quantity = self._get_column_value(row, 'min_quantity')
            
            # Calcola scorta minima se non specificata
            if not min_quantity:
                min_quantity = max(1, quantity // 4)  # 25% della quantità
            else:
                try:
                    min_quantity = int(float(min_quantity))
                except (ValueError, TypeError):
                    min_quantity = max(1, quantity // 4)
            
            # Rileva tipo vino se non specificato
            if not wine_type:
                wine_type = self._detect_wine_type(str(name))
            
            return {
                'name': str(name).strip(),
                'producer': str(producer).strip(),
                'quantity': quantity,
                'min_quantity': min_quantity,
                'vintage': int(vintage) if vintage else None,
                'wine_type': str(wine_type).strip() if wine_type else 'rosso',
                'region': str(region).strip() if region else None,
                'cost_price': float(price) if price else None,
                'notes': f"Importato da file"
            }
            
        except Exception as e:
            self.logger.error(f"Errore estrazione vino da riga: {e}")
            return None
    
    def _get_column_value(self, row: Dict[str, Any], key: str) -> Any:
        """
        Ottieni valore da una colonna
        
        Args:
            row: Riga del file
            key: Chiave da cercare
            
        Returns:
            Valore della colonna
        """
        # Cerca la chiave esatta
        if key in row:
            return row[key]
        
        # Cerca pattern comuni
        patterns = {
            'name': ['nome', 'name', 'etichetta', 'vino', 'wine', 'prodotto'],
            'producer': ['produttore', 'producer', 'marca', 'brand', 'casa'],
            'quantity': ['quantità', 'quantity', 'qty', 'stock', 'magazzino', 'bottiglie'],
            'vintage': ['annata', 'vintage', 'anno', 'year'],
            'wine_type': ['tipo', 'type', 'colore', 'color'],
            'region': ['regione', 'region', 'zona', 'area'],
            'price': ['prezzo', 'price', 'costo', 'cost'],
            'min_quantity': ['minimo', 'min', 'scorta_minima', 'min_stock']
        }
        
        if key in patterns:
            for pattern in patterns[key]:
                for col_name, value in row.items():
                    if pattern in col_name.lower():
                        return value
        
        return None
    
    
    def _detect_wine_type(self, name: str) -> str:
        """
        Rileva il tipo di vino dal nome
        
        Args:
            name: Nome del vino
            
        Returns:
            Tipo di vino
        """
        name_lower = name.lower()
        
        if any(word in name_lower for word in ['champagne', 'spumante', 'prosecco', 'franciacorta', 'cava']):
            return 'spumante'
        elif any(word in name_lower for word in ['bianco', 'white', 'chardonnay', 'pinot grigio', 'sauvignon']):
            return 'bianco'
        elif any(word in name_lower for word in ['rosato', 'rose', 'rosé', 'rosato']):
            return 'rosato'
        else:
            return 'rosso'  # Default
